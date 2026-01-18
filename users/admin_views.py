"""
Кастомные views для админ-панели
Дашборд, аналитика, экспорт данных
"""

from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse
from django.utils import timezone
from datetime import datetime, timedelta
import json

from booking.analytics import (
    get_financial_stats,
    get_occupancy_stats,
    get_clients_stats
)
from users.analytics import get_admin_dashboard_stats


@staff_member_required
def analytics_dashboard_view(request):
    """
    Главный дашборд с основными метриками
    """
    # Получаем параметры фильтрации
    period = request.GET.get('period', '30days')

    today = timezone.now().date()

    if period == '7days':
        start_date = today - timedelta(days=7)
    elif period == '30days':
        start_date = today - timedelta(days=30)
    elif period == '90days':
        start_date = today - timedelta(days=90)
    elif period == 'year':
        start_date = today - timedelta(days=365)
    elif period == 'custom':
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            today = end_date
        else:
            start_date = today - timedelta(days=30)
            end_date = today
    else:
        start_date = today - timedelta(days=30)

    end_date = today

    # Собираем все статистики
    try:
        financial = get_financial_stats(start_date, end_date)
        occupancy = get_occupancy_stats(start_date, end_date)
        clients = get_clients_stats(start_date, end_date)
        basic = get_admin_dashboard_stats()
    except Exception as e:
        # Если ошибка - показываем пустые данные
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading dashboard stats: {str(e)}", exc_info=True)

        financial = {'total_revenue': 0, 'paid_amount': 0, 'unpaid_amount': 0, 'monthly_revenue': [], 'revenue_by_court': []}
        occupancy = {'overall_occupancy_rate': 0, 'hourly_occupancy': [], 'weekday_occupancy': [], 'court_occupancy': []}
        clients = {'new_users': 0, 'active_users': 0, 'avg_ltv': 0, 'retention_rate': 0, 'top_clients': []}
        basic = {'today_bookings': 0, 'today_revenue': 0, 'occupancy_rate': 0, 'new_users_this_week': 0}

    context = {
        'financial': financial,
        'occupancy': occupancy,
        'clients': clients,
        'basic': basic,
        'period': period,
        'start_date': start_date,
        'end_date': end_date,

        # Сериализуем для Chart.js
        'financial_json': json.dumps(financial, default=str),
        'occupancy_json': json.dumps(occupancy, default=str),
        'clients_json': json.dumps(clients, default=str),
    }

    return render(request, 'admin_custom/analytics/dashboard.html', context)


@staff_member_required
def dashboard_stats_api(request):
    """
    API для получения статистики (для AJAX обновлений)
    """
    try:
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            today = timezone.now().date()
            start_date = today - timedelta(days=30)
            end_date = today

        stats = {
            'financial': get_financial_stats(start_date, end_date),
            'occupancy': get_occupancy_stats(start_date, end_date),
            'clients': get_clients_stats(start_date, end_date),
        }

        return JsonResponse(stats, safe=False)

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in dashboard_stats_api: {str(e)}", exc_info=True)

        return JsonResponse({
            'error': str(e)
        }, status=500)


@staff_member_required
def export_excel(request):
    """
    Экспорт данных в Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from io import BytesIO

        # Получаем параметры
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            today = timezone.now().date()
            start_date = today - timedelta(days=30)
            end_date = today

        # Получаем данные
        financial = get_financial_stats(start_date, end_date)
        occupancy = get_occupancy_stats(start_date, end_date)
        clients = get_clients_stats(start_date, end_date)

        # Создаем workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Аналитика"

        # Заголовок
        ws['A1'] = 'Paddle Booking - Отчет по аналитике'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
        ws.merge_cells('A2:D2')

        # Финансы
        row = 4
        ws[f'A{row}'] = 'ФИНАНСОВЫЕ ПОКАЗАТЕЛИ'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='9ef01a', end_color='9ef01a', fill_type='solid')
        row += 1

        ws[f'A{row}'] = 'Общий доход:'
        ws[f'B{row}'] = f"{financial['total_revenue']} ₽"
        row += 1

        ws[f'A{row}'] = 'Оплачено:'
        ws[f'B{row}'] = f"{financial['paid_amount']} ₽"
        row += 1

        ws[f'A{row}'] = 'Не оплачено:'
        ws[f'B{row}'] = f"{financial['unpaid_amount']} ₽"
        row += 1

        ws[f'A{row}'] = 'Прогноз на месяц:'
        ws[f'B{row}'] = f"{financial['forecast_next_month']} ₽"
        row += 2

        # Загруженность
        ws[f'A{row}'] = 'ЗАГРУЖЕННОСТЬ'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='9ef01a', end_color='9ef01a', fill_type='solid')
        row += 1

        ws[f'A{row}'] = 'Общая загруженность:'
        ws[f'B{row}'] = f"{occupancy['overall_occupancy_rate']}%"
        row += 1

        ws[f'A{row}'] = 'Всего бронирований:'
        ws[f'B{row}'] = occupancy['total_bookings']
        row += 1

        ws[f'A{row}'] = 'Забронировано часов:'
        ws[f'B{row}'] = occupancy['total_booked_hours']
        row += 2

        # Клиенты
        ws[f'A{row}'] = 'КЛИЕНТЫ'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='9ef01a', end_color='9ef01a', fill_type='solid')
        row += 1

        ws[f'A{row}'] = 'Новых пользователей:'
        ws[f'B{row}'] = clients['new_users']
        row += 1

        ws[f'A{row}'] = 'Активных пользователей:'
        ws[f'B{row}'] = clients['active_users']
        row += 1

        ws[f'A{row}'] = 'Средний LTV:'
        ws[f'B{row}'] = f"{clients['avg_ltv']} ₽"
        row += 1

        ws[f'A{row}'] = 'Retention rate:'
        ws[f'B{row}'] = f"{clients['retention_rate']}%"
        row += 2

        # Топ клиенты
        ws[f'A{row}'] = 'ТОП КЛИЕНТЫ'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        ws[f'A{row}'] = 'Имя'
        ws[f'B{row}'] = 'Бронирований'
        ws[f'C{row}'] = 'Потрачено'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'].font = Font(bold=True)
        row += 1

        for client in clients['top_clients'][:10]:
            ws[f'A{row}'] = client['name']
            ws[f'B{row}'] = client['bookings_count']
            ws[f'C{row}'] = f"{client['total_spent']} ₽"
            row += 1

        # Отдаем файл
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'paddle_analytics_{timezone.now().date()}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'

        return response

    except ImportError:
        return JsonResponse({
            'error': 'openpyxl не установлен. Выполните: pip install openpyxl'
        }, status=500)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error exporting to Excel: {str(e)}", exc_info=True)

        return JsonResponse({
            'error': f'Ошибка экспорта: {str(e)}'
        }, status=500)
